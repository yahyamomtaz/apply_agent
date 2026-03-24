import os
import re
import json
import argparse
from datetime import datetime
from typing import Optional, TypedDict, Annotated, Literal
from dotenv import load_dotenv
from bs4 import BeautifulSoup
from langchain_anthropic import ChatAnthropic
from langchain_core.messages import AnyMessage, SystemMessage, HumanMessage
from langchain_core.tools import tool
from langgraph.graph import START, StateGraph, END
from langgraph.graph.message import add_messages
from langgraph.prebuilt import ToolNode
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from anthropic import RateLimitError as AnthropicRateLimitError
from playwright.sync_api import sync_playwright
import PyPDF2
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import requests
import chromadb
from chromadb.utils import embedding_functions
from sentence_transformers import SentenceTransformer
import PyPDF2 as _PyPDF2

load_dotenv()

llm = ChatAnthropic(model="claude-sonnet-4-5", temperature=0.5, max_retries=0, max_tokens=4096)

def _make_retry(fn):
    return retry(
        retry=retry_if_exception_type(AnthropicRateLimitError),
        wait=wait_exponential(multiplier=1, min=5, max=60),
        stop=stop_after_attempt(5),
        reraise=True,
    )(fn)

#CHROMA DB
class _LocalEmbeddingFunction(embedding_functions.EmbeddingFunction):
    def __init__(self):
        print("📦 Loading local embedding model (all-MiniLM-L6-v2)...")
        self._model = SentenceTransformer("all-MiniLM-L6-v2")
    def __call__(self, input: list[str]) -> list[list[float]]:
        return self._model.encode(input, convert_to_numpy=True).tolist()

CHROMA_PATH = "./chroma_db"
COLLECTION_NAME = "phd_positions"

_chroma_client = chromadb.PersistentClient(path=CHROMA_PATH)

_collection = _chroma_client.get_or_create_collection(
    name=COLLECTION_NAME,
    embedding_function=_LocalEmbeddingFunction(),
    metadata={"hnsw:space": "cosine"}
)

# Shared State
# Flows through all nodes. Scraper fills the top half, Writer fills the bottom half, Supervisor reads everything.

class SharedState(TypedDict):
    # inputs
    url:                    Optional[str]
    cv_path:                Optional[str]
    output_filename:        Optional[str]
    no_cover_letter:        Optional[bool]

    # scraper outputs → writer inputs
    position_title:         Optional[str]
    position_university:    Optional[str]
    position_description:   Optional[str]
    position_requirements:  Optional[str]
    cv_text:                Optional[str]
    match_score:            Optional[int]
    match_verdict:          Optional[str]
    match_recommendation:   Optional[str]

    # writer outputs
    cover_letter_filename:  Optional[str]
    rag_stored:             Optional[bool]

# Scraper Tools

@tool
def fetch_webpage(url: str) -> str:
    """
    Fetch HTML content from a webpage using Playwright.

    Args:
        url: The URL to fetch

    Returns:
        The cleaned text content of the webpage
    """
    print(f"Fetching HTML from: {url}")
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            page.set_extra_http_headers({
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
            })
            page.goto(url, wait_until="networkidle", timeout=30000)
            html = page.content()
            browser.close()

            soup = BeautifulSoup(html, "html.parser")
            for tag in soup(["script", "style", "nav", "header", "footer", "aside", "form", "button"]):
                tag.decompose()

            for element in soup.find_all(class_=["navbar", "footer", "menu", "dropdown", "language-selector"]):
                element.decompose()

            text = soup.get_text(separator="\n", strip=True)
            cleaned = "\n".join([line for line in text.split("\n") if line.strip()])

            print(f"✅ Fetched and cleaned {len(cleaned)} characters")
            return cleaned[:60000]

    except Exception as e:
        print(f"ERROR in fetching webpage: {str(e)}")
        return {"error": str(e)}


@tool
def extract_cv_text(pdf_path: str) -> str:
    """
    Extract text from a CV PDF file.

    Args:
        pdf_path: The path to the CV PDF file.

    Returns:
        The extracted text from the CV PDF file.
    """
    try:
        with open(pdf_path, "rb") as pdf_file:
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"

        print(f"CV length: {len(text)}")
        return text.strip()

    except Exception as e:
        error_msg = f"Error extracting CV: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

@tool
def match_score_position(
    cv_text: str,
    position_title: str,
    position_description: str,
    position_requirements: str,
) -> str:
    """
    Score how well the applicant's CV matches a PhD position on a 0-100 scale.
    Returns a JSON string with the score, a short verdict, and per-category breakdown.

    Args:
        cv_text: Extracted text from the applicant's CV
        position_title: Title of the PhD position
        position_description: Full description of the position
        position_requirements: Required qualifications for the position

    Returns:
        JSON string with score, verdict, breakdown, and recommendation
    """
    print(f"Calculating match score for: {position_title}")
    prompt = f"""You are an expert academic recruiter. Score how well the applicant's CV matches the PhD position below.
    
    APPLICANT CV:{cv_text}

    POSITION TITLE:{position_title}

    POSITION REQUIREMENTS:{position_requirements}

    POSITION DESCRIPTION:{position_description}

    Respond with ONLY a valid JSON object — no markdown, no explanation outside the JSON.

Score each category from 0-100, then compute an overall weighted score:
- technical_skills (weight 30%): programming, tools, methods mentioned in requirements
- research_experience (weight 30%): publications, projects, lab experience
- domain_alignment (weight 25%): how closely the applicant's field matches the position's focus
- education (weight 15%): degree level, institution, relevant coursework

Return exactly this structure:
{{
  "overall_score": <0-100 integer>,
  "verdict": "<one of: Excellent Match | Good Match | Partial Match | Weak Match>",
  "recommendation": "<one sentence: should they apply, and why>",
  "breakdown": {{
    "technical_skills": <0-100>,
    "research_experience": <0-100>,
    "domain_alignment": <0-100>,
    "education": <0-100>
  }},
  "strengths": ["<strength 1>", "<strength 2>", "<strength 3>"],
  "gaps": ["<gap 1>", "<gap 2>"]
}}"""
    try:
        response = llm.invoke([HumanMessage(content=prompt)])
        raw = response.content.strip()

        # Strip markdown code fences if present
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip()

        # Validate it's proper JSON before returning
        parsed = json.loads(raw)
        score = parsed.get("overall_score", "N/A")
        verdict = parsed.get("verdict", "N/A")
        print(f"✅ Match score: {score}/100 — {verdict}")
        return json.dumps(parsed)

    except Exception as e:
        error_msg = f"Error scoring position: {str(e)}"
        print(f"❌ {error_msg}")
        return json.dumps({"overall_score": -1, "verdict": "Error", "recommendation": error_msg, "breakdown": {}, "strengths": [], "gaps": []})

@tool
def save_phd_position(
    title: str,
    university: str,
    deadline: str,
    description: str,
    requirements: str,
    links: str,
    match_score: int,
    match_verdict: str,
    match_recommendation: str
) -> str:
    """
    Save a PhD position to a local Excel spreadsheet database (phd_positions.xlsx).

    Args:
        title: Title of the PhD position
        university: University offering the position
        deadline: Application deadline
        description: Full description of the position
        requirements: Required qualifications
        links: Application links (comma-separated if multiple)
        match_score: Overall match score 0-100 (-1 if not scored)
        match_verdict: Verdict label (e.g. "Good Match")
        match_recommendation: One-sentence recommendation

    Returns:
        Success or error message
    """
    print("Saving to Excel spreadsheet...")

    FILENAME = "phd_positions.xlsx"
    HEADERS = ["Title", "University", "Deadline", "Description", "Requirements", "Links", "Match Score", "Verdict", "Recommendation"]
    new_row = [title, university, deadline, description, requirements, links, match_score, match_verdict, match_recommendation]

    try:
        if os.path.exists(FILENAME):
            wb = openpyxl.load_workbook(FILENAME)
            ws = wb.active

            # Check for duplicates (match on title + university)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == title and row[1] == university:
                    msg = "⚠️  Duplicate position found, skipping"
                    print(msg)
                    return msg

            ws.append(new_row)
            total = ws.max_row - 1  # subtract header row

        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "PhD Positions"

            # Write styled header row
            ws.append(HEADERS)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Set column widths
            col_widths = [40, 30, 15, 60, 60, 50]
            for i, width in enumerate(col_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

            ws.row_dimensions[1].height = 25

            ws.append(new_row)
            total = 1

        # Wrap text in all data cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(FILENAME)
        msg = f"✅ Saved! Total positions in database: {total}"
        print(msg)
        return msg

    except Exception as e:
        error_msg = f"Error saving data: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

# Writer Tools

@tool
def retrieve_similar_letters(
    position_title: str,
    position_description: str,
    n_results: int = 3
) -> str:
    """
    Retrieve the most similar past cover letters from the ChromaDB vector store.
    Uses the position title + description as the query to find semantically
    similar letters written for comparable roles.

    Args:
        position_title: Title of the current PhD position
        position_description: Description of the current position
        n_results: Number of similar letters to retrieve (default 3)

    Returns:
        A formatted string containing the retrieved letters and their metadata,
        or a message indicating no past letters exist yet.
    """
    print(f"🔍 Retrieving similar past cover letters for: {position_title}")
    try:
        total = _collection.count()
        if total == 0:
            print("ℹ️  No past cover letters in store yet — skipping retrieval")
            return "NO_HISTORY: No past cover letters found. Generate without examples."

        n_results = min(n_results, total)
        query_text = f"{position_title}\n{position_description}"
        
        results = _collection.query(
            query_texts=[query_text],
            n_results=n_results,
            include=["documents", "metadatas", "distances"]
        )

        documents = results["documents"][0]
        metadatas = results["metadatas"][0]
        distances = results["distances"][0]

        if not documents:
            return "NO_HISTORY: No similar cover letters found."

        output_parts = [
            f"Retrieved {len(documents)} similar past cover letter(s) "
            f"from {total} total in the store:\n"
        ]

        for i, (doc, meta, dist) in enumerate(zip(documents, metadatas, distances), 1):
            similarity = round((1 - dist) * 100, 1)
            print("similarity: ", similarity)

            output_parts.append(
                f"--- EXAMPLE {i} ---\n"
                f"Position : {meta.get('position_title', 'N/A')}\n"
                f"University: {meta.get('university', 'N/A')}\n"
                f"Score     : {meta.get('match_score', 'N/A')}/100\n"
                f"Similarity: {similarity}%\n"
                f"Date      : {meta.get('date', 'N/A')}\n\n"
                f"{doc}\n"
            )

        retrieved = "\n".join(output_parts)
        print(f"✅ Retrieved {len(documents)} similar letter(s)")
        return retrieved

    except Exception as e:
        error_msg = f"Error retrieving similar letters: {str(e)}"
        print(f"❌ {error_msg}")
        return f"RETRIEVAL_ERROR: {error_msg}"


@tool
def store_cover_letter_in_rag(
    cover_letter_text: str,
    position_title: str,
    university: str,
    match_score: int = -1
) -> str:
    """
    Store a generated cover letter in the ChromaDB vector store for future RAG retrieval.
    Call this AFTER saving the cover letter to a .txt file.

    Args:
        cover_letter_text: The final humanized cover letter text
        position_title: Title of the PhD position this letter was written for
        university: University of the position
        match_score: Match score for this position (0-100, -1 if unknown)

    Returns:
        Success or error message
    """
    print(f"📥 Storing cover letter in RAG for: {position_title}")
    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = position_title.lower().replace(" ", "_")[:40]
        safe_uni = university.lower().replace(" ", "_")[:20]
        doc_id = f"{safe_title}__{safe_uni}__{timestamp}"

        metadata = {
            "position_title": position_title,
            "university": university,
            "match_score": match_score,
            "date": datetime.now().strftime("%Y-%m-%d")
        }

        _collection.add(
            documents=[cover_letter_text],
            metadatas=[metadata],
            ids=[doc_id]
        )
        
        total = _collection.count()
        msg = f"✅ Stored in RAG vector store (total letters: {total})"
        print(msg)
        return msg

    except Exception as e:
        error_msg = f"Error storing in RAG: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

@tool
def generate_cover_letter(
    cv_text: str,
    position_title: str,
    position_description: str
) -> str:
    """
    Generate a tailored "My Interest in the Role" section for a PhD position
    based on the applicant's CV and the position description.

    Args:
        cv_text: Text extracted from the applicant's CV
        position_title: Title of the PhD position
        position_description: Full description of the position

    Returns:
        A tailored 2-3 paragraph interest section as a string
    """
    print(f"Generating cover letter section for: {position_title}")

    prompt = f"""You MUST write EXACTLY 2-3 paragraphs for the "My Interest in the Role" section of a PhD cover letter.

APPLICANT CV:
{cv_text}

POSITION:
- Title: {position_title}
- Description: {position_description}

CRITICAL REQUIREMENTS - YOU MUST FOLLOW THESE:
✓ Output EXACTLY 2-3 paragraphs of body text
✓ NO headers (like "My Interest in the Role")
✓ NO titles or section labels
✓ NO greetings ("Dear...") or sign-offs ("Sincerely...")
✓ NO salutations or closings
✓ ONLY the paragraph content for the interest section
✓ Start directly with the first paragraph
✓ End with the last paragraph

FORBIDDEN - Do NOT include:
❌ Any other sections of a cover letter
❌ Contact information
❌ Dates
❌ Recipient information
❌ Phrases like: "I am excited to", "leverage", "I'm particularly drawn", "I am writing to"

WRITING STYLE:
- Highlight how the applicant's qualifications match the PhD position's requirements
- Show experience as an AI Researcher — include specific achievements such as publications and research contributions from the CV
- Show experience on research and publishing papers
- Explain how the applicant's skills directly address the key requirements in the position description
- Write a compelling paragraph illustrating how the applicant's experience with specific technologies, tools, or methods can contribute to the position's current projects or goals
- Write authentically in the first person as a PhD candidate
- Be specific, avoid generic statements"""

    try:
        response = llm.invoke([HumanMessage(content=prompt)])
        cover_letter = response.content

        print(f"✅ Generated cover letter section ({len(cover_letter)} characters)")
        return cover_letter

    except Exception as e:
        error_msg = f"Error generating cover letter: {str(e)}"
        print(f"❌ {error_msg}")
        return error_msg

@tool
def humanize_cover_letter(text: str) -> str:
    """
    Humanize the generated cover letter section using the Rephrasy API
    to make it sound more natural and less AI-generated.

    Args:
        text: The cover letter text to humanize

    Returns:
        Humanized version of the text
    """
    print("🔥 Humanizing cover letter...")
    try:
        response = requests.post(
            "https://v2-humanizer.rephrasy.ai/api",
            headers={"Authorization": f"Bearer {os.getenv('REPHRASY_API_KEY')}"},
            json={"text": text, "model": "v3", "style": "professional"}
        )
        response.raise_for_status()
        result = response.json().get("result", text)
        print(f"✅ Humanized ({len(result)} characters)")
        return result

    except Exception as e:
        print(f"⚠️  Humanizer failed ({str(e)}), returning original text")
        return text


@tool
def save_cover_letter(interest_section: str, filename: str) -> str:
    """
    Save the generated "My Interest in the Role" section to a text file.

    Args:
        interest_section: The 2-3 paragraphs generated for the interest section
        filename: Name of the output file (will add .txt if not present)

    Returns:
        Success or error message
    """
    try:
        if not filename.endswith('.txt'):
            filename += '.txt'

        with open(filename, 'w', encoding='utf-8') as f:
            f.write(interest_section)

        print(f"✅ Cover letter section saved to: {filename}")
        return f"✅ Cover letter section saved to: {filename}"

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return f"Error: {str(e)}"


SCRAPER_TOOLS = [
    fetch_webpage,
    save_phd_position,
    extract_cv_text,
    match_score_position,
]

WRITER_TOOLS = [
    generate_cover_letter,
    humanize_cover_letter,
    retrieve_similar_letters,
    store_cover_letter_in_rag,
    save_cover_letter
]

scraper_llm = llm.bind_tools(SCRAPER_TOOLS, parallel_tool_calls=False)
writer_llm = llm.bind_tools(WRITER_TOOLS, parallel_tool_calls=False)

@_make_retry
def _call_scraper(message): return scraper_llm.invoke(message)

@_make_retry
def _call_writer(message): return writer_llm.invoke(message)

SCRAPER_SYSTEM = """You are the Scraper Agent in a PhD application pipeline.

YOUR ONLY JOB: gather position information and score how well the CV matches it.

TOOLS:
- fetch_webpage(url): scrape the position page
- extract_cv_text(pdf_path): read the CV
- match_score_position(cv_text, position_title, position_description, position_requirements): score the match
- save_phd_position(...): save position + score to Excel

WORKFLOW — exact order, no skipping:
1. fetch_webpage(url)
2. extract_cv_text(cv_path)
3. match_score_position(cv_text, title, description, requirements) — extract COMPLETE text, never summarize
4. save_phd_position(title, university, deadline, description, requirements, links, match_score, match_verdict, match_recommendation)

When all 4 steps are done, respond with ONLY this JSON (no other text):
{
  "position_title": "<title>",
  "position_university": "<university>",
  "position_description": "<full description>",
  "position_requirements": "<full requirements>",
  "match_score": <0-100 integer>,
  "match_verdict": "<Excellent Match|Good Match|Partial Match|Weak Match>",
  "match_recommendation": "<one sentence>"
}"""

def scraper_agent(state: SharedState) -> SharedState:
    print("\n" + "─" * 60)
    print("🕷️  SCRAPER AGENT")
    print("─" * 60)

    messages = [
        SystemMessage(content=SCRAPER_SYSTEM),
        HumanMessage(content=f"""URL: {state['url']}\nCV path: {state['cv_path']}""")
    ]

    while True:
        response = _call_scraper(messages)
        messages.append(response)
        if not getattr(response, "tool_calls", None):
            break
        tool_results = ToolNode(SCRAPER_TOOLS).invoke({"messages": messages})
        messages.extend(tool_results["messages"])

    raw = response.content
    if isinstance(raw, list):
        raw = " ".join(b.get("text", "") if isinstance(b, dict) else str(b) for b in raw)

    try:
        clean = raw.strip()
        if clean.startswith("```"):
            clean = clean.split("```")[1]
            if clean.startswith("json"):
                clean = clean[4:]
        result = json.loads(clean.strip())
        print(f"✅ Scraper done — {result.get('match_score')}/100 ({result.get('match_verdict')})")
        return {
            **state,
            "position_title":         result.get("position_title", ""),
            "position_university":    result.get("position_university", ""),
            "position_description":   result.get("position_description", ""),
            "position_requirements":  result.get("position_requirements", ""),
            "match_score":            int(result.get("match_score", -1)),
            "match_verdict":          result.get("match_verdict", ""),
            "match_recommendation":   result.get("match_recommendation", ""),
        }
    except Exception as e:
        print(f"❌ Scraper parse error: {e}")
        return {**state, "match_score": -1} 
    

# Writer Agent

WRITER_SYSTEM = """You are the Writer Agent in a PhD application pipeline.

YOUR ONLY JOB: generate the best possible tailored cover letter and store it.

TOOLS:
- retrieve_similar_letters(position_title, position_description): get past examples from RAG
- generate_cover_letter(cv_text, position_title, position_description, similar_letters): generate section
- humanize_cover_letter(text): make text more natural
- save_cover_letter(interest_section, filename): save to .txt — use EXACTLY the filename given to you
- store_cover_letter_in_rag(cover_letter_text, position_title, university, match_score): store for future RAG

WORKFLOW — exact order:
1. retrieve_similar_letters(position_title, position_description)
2. generate_cover_letter(cv_text, position_title, position_description, similar_letters=<step 1 output>)
3. humanize_cover_letter(text=<step 2 output>)
4. save_cover_letter(interest_section=<step 3 output>, filename=<exactly the filename given>)
5. store_cover_letter_in_rag(cover_letter_text=<step 3 output>, position_title, university, match_score)"""

def writer_agent(state: SharedState) -> SharedState:
    print("\n" + "─" * 60)
    print("✍️  WRITER AGENT")
    print("─" * 60)

    messages = [
        SystemMessage(content=WRITER_SYSTEM),
        HumanMessage(content=f"""Position Title: {state['position_title']}
University: {state['position_university']}
Match Score: {state['match_score']}/100 ({state['match_verdict']})
Output filename: {state['output_filename']}

POSITION DESCRIPTION:
{state['position_description']}

CV TEXT:
{state['cv_text']}""")
    ]

    while True:
        response = _call_writer(messages)
        messages.append(response)
        if not getattr(response, "tool_calls", None):
            break
        tool_results = ToolNode(WRITER_TOOLS).invoke({"messages": messages})
        messages.extend(tool_results["messages"])

    print(f"✅ Writer done")
    return {**state, "cover_letter_filename": state["output_filename"], "rag_stored": True}

# Supervisor Agent

def supervisor(state: SharedState) -> SharedState:
    """Entry point — pre-loads CV text into state, then routes."""
    print("\n" + "═" * 60)
    if state.get("rewrite_mode"):
        print("🎯 SUPERVISOR — rewrite mode, dispatching to Rewriter Agent")
        print("═" * 60)
        return state

    print("🎯 SUPERVISOR — dispatching to Scraper Agent")
    print("═" * 60)

    # Pre-extract CV text here so it never travels through JSON
    # (embedding raw CV text in a JSON string breaks json.loads on newlines/special chars)
    cv_text = ""
    if state.get("cv_path"):
        try:
            with open(state["cv_path"], "rb") as f:
                reader = _PyPDF2.PdfReader(f)
                cv_text = "".join(p.extract_text() + "\n" for p in reader.pages).strip()
            print(f"📄 Supervisor pre-loaded CV ({len(cv_text)} chars)")
        except Exception as e:
            print(f"⚠️  Supervisor CV pre-load failed: {e}")

    return {**state, "cv_text": cv_text}

def route_after_scraper(state: SharedState) -> Literal["writer_agent", "end_node"]:
    """After scraper runs: go to writer if score ≥ 50 and cover letter not disabled."""
    score = state.get("match_score", -1)
    no_cl = state.get("no_cover_letter", False)

    if score >= 50 and not no_cl:
        print(f"\n🎯 SUPERVISOR — score {score}/100 ≥ 50, dispatching to Writer Agent")
        return "writer_agent"
    else:
        reason = "--no-cover-letter flag" if no_cl else f"score {score}/100 < 50"
        print(f"\n🎯 SUPERVISOR — skipping Writer Agent ({reason})")
        return "end_node"

def end_node(state: SharedState) -> SharedState:
    """Print final summary."""
    print("\n" + "═" * 60)
    print("📋 FINAL SUMMARY")
    print("═" * 60)
    print(f"Position  : {state.get('position_title', 'N/A')}")
    print(f"University: {state.get('position_university', 'N/A')}")
    print(f"Score     : {state.get('match_score', 'N/A')}/100 — {state.get('match_verdict', 'N/A')}")
    print(f"Rec.      : {state.get('match_recommendation', 'N/A')}")
    if state.get("cover_letter_filename"):
        print(f"Cover letter: ✅ {state['cover_letter_filename']}")
        print(f"RAG stored  : {'✅' if state.get('rag_stored') else '❌'}")
    else:
        score = state.get("match_score", -1)
        reason = "score below 50" if score < 50 else "--no-cover-letter flag"
        print(f"Cover letter: ⚠️  skipped ({reason})")
    print("═" * 60 + "\n")
    return state


# Graph

builder = StateGraph(SharedState)

builder.add_node("supervisor", supervisor)
builder.add_node("scraper_agent", scraper_agent)
builder.add_node("writer_agent", writer_agent)
builder.add_node("end_node", end_node)

builder.add_edge(START, "supervisor")
builder.add_edge("supervisor", "scraper_agent")
builder.add_conditional_edges("scraper_agent", route_after_scraper, {"writer_agent": "writer_agent", "end_node": "end_node"})
builder.add_edge("writer_agent", "end_node")
builder.add_edge("end_node", END)

graph = builder.compile()


def main():
    parser = argparse.ArgumentParser(
        description="PhD Application Assistant — Multi-Agent (Supervisor + Scraper + Writer)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python phd_agent.py --url "https://academicpositions.com/ad/..."
  python phd_agent.py --url "https://..." --cv my_cv.pdf
  python phd_agent.py --url "https://..." --no-cover-letter
  python phd_agent.py --url "https://..." --output cover_letter_eth.txt
        """
    )
    parser.add_argument("--url",             required=True, help="URL of the PhD position")
    parser.add_argument("--cv",              default="templates/cv.pdf", help="Path to CV PDF")
    parser.add_argument("--no-cover-letter", action="store_true", help="Only scrape and score")
    parser.add_argument("--output",          help="Custom cover letter filename")
    args = parser.parse_args()

    if not os.path.exists(args.cv):
        print(f"❌ CV not found at '{args.cv}'. Use --cv to specify the path.")
        return

    if args.output:
        output_filename = args.output
    else:
        slug = args.url.rstrip("/").split("/")[-1]
        slug = re.sub(r"[^a-z0-9]+", "_", slug.lower()).strip("_")
        output_filename = f"{slug}_cover_letter.txt"

    print("\n" + "═" * 80)
    print("PhD APPLICATION ASSISTANT  —  Multi-Agent")
    print("═" * 80)
    print(f"URL       : {args.url}")
    print(f"CV        : {args.cv}")
    print(f"Mode      : {'Scrape + Score only' if args.no_cover_letter else 'Scrape + Score + Generate'}")
    print(f"RAG store : {_collection.count()} past letter(s)")
    print(f"Output    : {output_filename}")
    print("═" * 80 + "\n")

    try:
        graph.invoke({
            "url":                  args.url,
            "cv_path":              args.cv,
            "output_filename":      output_filename,
            "no_cover_letter":      args.no_cover_letter,
            "position_title":       None,
            "position_university":  None,
            "position_description": None,
            "position_requirements":None,
            "cv_text":              None,
            "match_score":          None,
            "match_verdict":        None,
            "match_recommendation": None,
            "cover_letter_filename":None,
            "rag_stored":           None,
        })
    except Exception as e:
        print(f"\n❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
